# IMPORT LIBRARIES
import os
import pandas as pd
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font, Protection
import sys
import tempfile
import shutil
import webbrowser
import json
import requests


# FUNCTIONS TO MANAGE FILES AND FOLDERS

def working_directory():
    """
    Returns the path to the project's parent directory. If the script is bundled into an executable, it returns the executable's directory.

    Args:
        None

    Returns:
        str: Path to the parent directory of the project or the directory of the executable if running as a bundled application.
    """
    if hasattr(sys, 'frozen'):
        # If the script is bundled into an executable, it returns the directory of the executable.
        directory = os.path.dirname(sys.executable)
    else:
        # If running as a normal script, it returns the parent directory of the project
        directory = os.path.abspath(os.path.join(__file__, *(['..'] * 3)))
    
    return directory  # Returns the directory



def project_directory():
    """
    Retrieves the path to the project directory, adjusting based on the execution context.

    Args:
        None

    Returns:
        str: The path to the project directory. If running as a bundled executable, returns the temporary directory; otherwise, returns the actual project directory path.
    """
    if hasattr(sys, '_MEIPASS'):
        # If running as a bundled executable, use the temporary directory
        project_directory = sys._MEIPASS
    else:
        # If running as a normal script, get the project directory
        project_directory = os.path.abspath(os.path.join(__file__, *(['..'] * 2)))
    
    return project_directory  # Return the project directory



def get_filename_or_directory():
    """
    Retrieves the name of the executable file if bundled, or the project directory if running as a script.

    Args:
        None

    Returns:
        str: Name of the executable file or project directory, depending on the execution context.
    """
    if getattr(sys, 'frozen', False):
        # If the script is bundled into an executable
        return os.path.basename(sys.executable)
    else:
        # If running as a normal script gets the directory name
        return os.path.basename(project_directory()).rstrip("/").rstrip("\\")
    
    
# FUNCTIONS INTO SCRIPTS
def choose_language():
    """
    Retrieves the language from the command line arguments. Defaults to English if no argument is provided.

    Args:
        None

    Returns:
        str: The selected language. Default is "en" if no language argument is provided.
    """
    # Get the language from the command line arguments
    language = "en"  # Default language
    if len(sys.argv) > 1:
        language = sys.argv[1]  # Take the argument if provided
    
    return language




def get_instructions(language):
    """
    Opens an Excel file, displays notes in column E, unlocks specified rows, and applies sheet protection.

    Args:
        excel_file_path (str): Path to the Excel file to be modified.
        additional_rows_to_unlock (int): Additional number of rows to unlock beyond those in the DataFrame.
        notes_title (str): Title for the notes in cell E2.
        excel_note1 (str): First note to display in cell E3.
        excel_note2 (str): Second note to display in cell E4.
        excel_note3 (str): Third note to display in cell E5.
        excel_note4 (str): Fourth note to display in cell E6.
        excel_note5 (str): Fifth note to display in cell E7.

    Returns:
        None: Saves the modified Excel file with notes and protection settings.

    Raises:
        FileNotFoundError: If the specified Excel file path does not exist.
        PermissionError: If there is an issue with file access or write permissions.
    """
    # Get the base directory of the project
    project_directory_path = project_directory()

    # Check if we're in a bundled (PyInstaller) environment
    if hasattr(sys, '_MEIPASS'):
        # Use the temporary directory to copy the files
        temp_dir = tempfile.mkdtemp()

        # Path to the instructions HTML file in the specified language
        instructions_path = os.path.join(sys._MEIPASS, 'instructions', f'Instructions - {language}.html')
        # If the file in the desired language is not found, use the English file
        if not os.path.exists(instructions_path):
            instructions_path = os.path.join(sys._MEIPASS, 'instructions', 'Instructions - en.html')

        # Copy the HTML file and CSS to the temporary directory
        shutil.copy(instructions_path, temp_dir)
        css_path = os.path.join(sys._MEIPASS, 'instructions', 'styles', 'styles.css')
        if os.path.exists(css_path):
            os.makedirs(os.path.join(temp_dir, 'styles'), exist_ok=True)
            shutil.copy(css_path, os.path.join(temp_dir, 'styles'))

        # Copy all images to the temporary directory
        pictures_path = os.path.join(sys._MEIPASS, 'instructions', 'pictures')
        temp_pictures_dir = os.path.join(temp_dir, 'pictures')
        os.makedirs(temp_pictures_dir, exist_ok=True)
        for image_name in [
            '1 - Move to folder.png',
            '2 - Main menu.png',
            '3 - excel template.png',
            '4 - Flash fill excel.gif',
            '5 - renaming.png',
            '6 - options to modify names.png'
        ]:
            image_path = os.path.join(pictures_path, image_name)
            if os.path.exists(image_path):
                shutil.copy(image_path, temp_pictures_dir)

        # New path for the HTML file in the temporary directory
        instructions_path = os.path.join(temp_dir, os.path.basename(instructions_path))
    
    else:
        # In development environment, the path is relative to the project directory
        instructions_path = os.path.join(project_directory_path, 'instructions', f'Instructions - {language}.html')
        # If the file in the desired language is not found, use the English file
        if not os.path.exists(instructions_path):
            instructions_path = os.path.join(project_directory_path, 'instructions', 'Instructions - en.html')

    # Try to open the file in the browser
    try:
        # Open the file in the default web browser
        webbrowser.open(f'file:///{os.path.abspath(instructions_path)}')

        # Return the content of the file
        with open(instructions_path, 'r', encoding='utf-8') as file:
            return file.read()  # Return the content of the file
    except FileNotFoundError:
        raise FileNotFoundError(f"Instructions file not found: {instructions_path}")
    except Exception as e:
        raise Exception(f"Error opening instructions file: {e}")




    
    
    
        
def load_available_languages():
    """
    Loads the available languages from a JSON file located in the 'locales' folder of the project's.
    
    Args:
        None

    Returns:
        dict: A dictionary containing the available languages and their corresponding details.
    """
    # Get the base directory of the project
    project_directory_path = project_directory() 
    # Load the languages dictionary from the JSON file
    json_path = os.path.join(project_directory_path, 'locales', 'languages.json')
    with open(json_path, 'r', encoding='utf-8') as file:
        return json.load(file)
    
    
    
    
def load_translations(language):
    """
    Loads translations from a JSON file based on the specified language.

    Args:
        language (str): The language code (e.g., 'es' for Spanish).

    Returns:
        dict: A dictionary with the translations for the specified language.
    """
    # Get the base directory of the project
    project_directory_path = project_directory() 
    # Build the path to the language-specific JSON file
    language_json_path = os.path.join(project_directory_path, 'locales', f'{language}.json')
    english_json_path = os.path.join(project_directory_path, 'locales', 'en.json')

    # Load the translations for the specified language
    translations = {}
    
    # Attempt to load the language file
    if os.path.exists(language_json_path):
        with open(language_json_path, 'r', encoding='utf-8') as file:
            translations = json.load(file)
    
    # Load the default English translations (en.json)
    with open(english_json_path, 'r', encoding='utf-8') as file:
        english_translations = json.load(file)

    # Replace missing translations with the English defaults
    for key in english_translations:
        if key not in translations:
            translations[key] = english_translations[key]

    return translations




def ask_replace(file_already_exist_title, file_already_exist_message):
    """
    If a file with the same name exists, asks if the user wants to replace it (using tkinter).

    Args:
        file_already_exist_title (str): The title of the message box.
        file_already_exist_message (str): The message displayed in the message box.

    Returns:
        bool: True if the user selects "Yes", False if the user selects "No".
    """
    root = tk.Tk()
    root.withdraw()  # Hides the main window
    return messagebox.askyesno(file_already_exist_title, file_already_exist_message)



def show_error(error_title, error_message):
    """
    Displays an error message using tkinter.

    Args:
        error_title (str): The title of the error message box.
        error_message (str): The error message displayed in the message box.

    Returns:
        None
    """
    root = tk.Tk()
    root.withdraw()  # Hides the main window
    messagebox.showerror(error_title, error_message)
    

def show_message(title, message):
    """
    Displays a message in a popup window using tkinter. If the message is a DataFrame, it converts it to a string.

    Args:
        title (str): The title of the popup window.
        message (str or pd.DataFrame): The message to display. If a DataFrame is provided, it will be converted to a string.

    Returns:
        None: This function does not return anything.

    Raises:
        None: This function does not raise any exceptions.
    """
    # Create a hidden window
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    # Convert the message to a string if it is a DataFrame
    if isinstance(message, pd.DataFrame):
        message = message.to_string()  # Convert DataFrame to string
    # Display the message in a popup window
    messagebox.showinfo(title, message)
    # Close the window after the message is closed
    root.destroy()




def modify_excel_dataframe(excel_file_path, additional_rows_to_unlock, notes_title, excel_note1, excel_note2, excel_note3, excel_note4, excel_note5):
    """
    Opens an Excel file, displays notes in column E, unlocks specified rows, and applies sheet protection.

    Args:
        excel_file_path (str): The path to the Excel file to modify.
        additional_rows_to_unlock (int): The number of additional rows to unlock beyond the DataFrame's rows.
        notes_title (str): The title note to display in cell E2.
        excel_note1 (str): The note to display in cell E3.
        excel_note2 (str): The note to display in cell E4.
        excel_note3 (str): The note to display in cell E5.
        excel_note4 (str): The note to display in cell E6.
        excel_note5 (str): The note to display in cell E7.

    Returns:
        None: This function does not return anything.
    """
    # Open the Excel file to display a note in cell E2
    wb = load_workbook(excel_file_path)
    ws = wb.active  # Select the first worksheet (active sheet)

    # Write notes in column E
    ws['E2'] = notes_title
    ws['E2'].font = Font(bold=True)  # Apply bold formatting
    ws['E3'] = excel_note1
    ws['E4'] = excel_note2
    ws['E5'] = excel_note3
    ws['E6'] = excel_note4
    ws['E7'] = excel_note5

    # Sheet protection
    # 1,048,576 rows is the default Excel limit. Ensure the sheet is unlocked without exceeding the limit.
    if ws.max_row >= 1048576 - additional_rows_to_unlock:
        ws.max_row = 1048576 - additional_rows_to_unlock
    # Unlock the rows in the DataFrame + additional rows
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row + additional_rows_to_unlock, min_col=1, max_col=16):
        for cell in row:
            cell.protection = Protection(locked=False)
    
    # Lock cells A1, B1, and C1
    for cell in ['A1', 'B1', 'C1']:
        ws[cell].protection = Protection(locked=True)

    # Enable sheet protection
    ws.protection.sheet = True

    # Allow specific actions on the sheet
    ws.protection.sort = False           # Allow sorting
    ws.protection.autoFilter = False     # Allow autofilters
    ws.protection.insertRows = False     # Allow row insertion
    ws.protection.deleteRows = False     # Allow row deletion
    ws.protection.insertColumns = False  # Allow column insertion
    ws.protection.deleteColumns = False  # Allow column deletion
    ws.protection.formatColumns = False  # Allow column formatting
    ws.protection.formatRows = False     # Allow row formatting

    # Save the modified file
    wb.save(excel_file_path)
    wb.close()

    
    

def unlock_excel_sheet(excel_path):
    """
    Disables the protection of an Excel sheet.

    Args:
        excel_path (str): The path to the Excel file to be modified.

    Returns:
        None
    """
    # Abrir el archivo Excel para mostrar una nota en la celda E2
    wb = load_workbook(excel_path)
    ws = wb.active  # Selecciona la primera hoja de trabajo (la activa)

    # Deshabilitar la protección de la hoja
    ws.protection.sheet = False

    # Guardar el archivo modificado
    wb.save(excel_path)
    wb.close()
  

  
def delete_extention(filename):
    """
    Removes the extension from a given filename.

    Args:
        filename (str): The name of the file including its extension.

    Returns:
        str: The filename without its extension.
    """
    # Find the position of the last dot
    dot_index = filename.rfind('.')
    
    # If there is no dot in the string, return the full string
    if dot_index == -1:
        return filename
    else:
        # Return the string from the beginning up to the last dot
        return filename[:dot_index]





def show_options(title, option1, option2, border1, border2, cancel_title):
    """
    Displays an options window with two buttons.
    Upon selecting an option, returns the number of the selected option (1 or 2).
    If canceled, closes only the options window without affecting other windows.

    Args:
        title (str): The title of the options window.
        option1 (str): The text for the first option button.
        option2 (str): The text for the second option button.
        border1 (str): The color of the border for the first option frame.
        border2 (str): The color of the border for the second option frame.
        cancel_title (str): The text for the cancel button.

    Returns:
        int: The selected option (1 or 2) or None if canceled.
    """
    def cancel_current_operation():
        """
        This sub-function cancels the operation and closes only the current options window.

        Args:
            None

        Returns:
            None
        """
        global options_window
        if options_window:
            options_window.destroy()  # Closes the options window
        
    def select_option(option_number):
        """
        Sub-function to handle option selection.

        Args:
            option_number (int): The number of the selected option (1 or 2).

        Returns:
            None
        """
        nonlocal result
        result = option_number
        cancel_current_operation()
        options_window.quit()  # Closes the options window

    global options_window
    options_window = tk.Toplevel()  # Create a secondary window
    options_window.title(title)  # Set the window title
    
    # Add the icon to the window
    options_window.iconbitmap(relative_route_to_file("assets", "icon.ico"))

    # Variable to store the result of the selection
    result = None

    # Handle the close (X) button of the window
    options_window.protocol("WM_DELETE_WINDOW", cancel_current_operation)

    # Frame for the first option
    frame1 = tk.Frame(options_window, highlightbackground=border1, highlightthickness=2)
    frame1.pack(pady=20, padx=20)
    
    button1 = tk.Button(frame1, text=option1, font=("Arial", 10), borderwidth=0, relief="flat",
                        bg="white", activebackground="lightgrey", 
                        padx=15, pady=10,  # Padding
                        command=lambda: select_option(1))
    button1.pack()

    # Frame for the second option
    frame2 = tk.Frame(options_window, highlightbackground=border2, highlightthickness=2)
    frame2.pack(pady=20, padx=20)

    button2 = tk.Button(frame2, text=option2, font=("Arial", 10), borderwidth=0, relief="flat",
                        bg="white", activebackground="lightgrey", 
                        padx=15, pady=10,  # Padding
                        command=lambda: select_option(2))
    button2.pack()

    # Cancel button
    cancel_button = tk.Button(options_window, text=cancel_title, font=("Arial", 10), command=cancel_current_operation, 
                               padx=10, pady=5, bg="red", fg="white")
    cancel_button.pack(side=tk.BOTTOM, anchor=tk.SE, padx=10, pady=10)  # Bottom-right corner

    # Keep the window open until an option is selected or the window is closed
    options_window.mainloop()

    # Return the result after completion
    return result







def copy_files_with_new_names(dataframe_to_process, source_folder, destination_folder, excel_error_file_doesnt_found, excel_error_file_already_proceced, excel_error_file_already_exist, excel_template_error_doesnt_found, excel_template_error_not_allowed, excel_column_status):
    """
    This function creates a new folder (if it doesn't exist), checks if there are files with the new names (if they exist, they will be marked with statuses). 
    It also checks that there are no duplicates in the new names (it will also mark them with statuses). 
    Finally, it copies the files from the source folder and pastes them with the new names in the destination folder. 
    All actions are logged in statuses and returned in the DataFrame.

    Args:
        dataframe_to_process (pd.DataFrame): The DataFrame containing the files or folders to be processed.
        source_folder (str): The source folder where the original files or folders are located.
        destination_folder (str): The destination folder where the renamed files or folders will be copied.
        excel_error_file_doesnt_found (str): Error message to indicate that a file was not found in the source.
        excel_error_file_already_proceced (str): Error message for already processed files.
        excel_error_file_already_exist (str): Error message for files that already exist in the destination.
        excel_template_error_doesnt_found (str): Error message for a template file not found.
        excel_template_error_not_allowed (str): Error message for a permission error when copying a file.
        excel_column_status (str): The column name in the DataFrame where the status will be stored.

    Returns:
        pd.DataFrame: The updated DataFrame with a new column containing the statuses of each operation.
    """
    # Create the destination folder if it doesn't exist
    if not os.path.exists(destination_folder):
        os.makedirs(destination_folder)

    # List to store statuses
    statuses = []
    processed_names = set()  # Set to track unique names for 'New full name'

    # Iterate over the DataFrame to copy files and folders
    for index, row in dataframe_to_process.iterrows():
        original_full_name = row.iloc[3]  # Column 4 = Original file or folder name
        new_full_name = row.iloc[4]  # Column 5 = New name for the file or folder in the destination

        # Full path for the source file or folder
        source_path = os.path.join(source_folder, original_full_name)
        # Full path for the destination file or folder (with "Modified name")
        destination_path = os.path.join(destination_folder, new_full_name)

        # Check if the file or folder exists in the source folder
        if not os.path.exists(source_path):
            statuses.append(excel_error_file_doesnt_found + " " + new_full_name)
            continue  # Skip if it doesn't exist in source

        # Check if the new name has already been processed (duplicate in DataFrame)
        if new_full_name in processed_names:
            statuses.append(excel_error_file_already_proceced)
            continue
        
        # Add the new name to the processed set
        processed_names.add(new_full_name)

        # Check if the file or folder already exists in the destination folder
        if os.path.exists(destination_path):
            # If it already exists, log the status and continue
            statuses.append(excel_error_file_already_exist)
            continue

        # Try to copy the file or folder
        try:
            if os.path.isdir(source_path):
                # If it's a folder, copy the entire folder structure
                shutil.copytree(source_path, destination_path)
            else:
                # If it's a file, only copy the file
                shutil.copy(source_path, destination_path)

            # Add 'Ok' status if the copy was successful
            statuses.append("Ok")
        except FileNotFoundError:
            # Handle the case where the file is not found
            statuses.append(excel_template_error_doesnt_found)
        except PermissionError:
            # Handle permission errors
            statuses.append(excel_template_error_not_allowed)
        except Exception as e:
            # Handle other errors during the copy process
            statuses.append(f"Error: {e}")

    # Add the list of statuses to the DataFrame as a new column
    dataframe_to_process[excel_column_status] = statuses

    # Return the DataFrame with the new column
    return dataframe_to_process




def ask_to_proceed(start_title, start_message):
    """
    This function prompts the user with a message box asking if they wish to continue with the file renaming process.
    
    Args:
        start_title (str): The title of the message box.
        start_message (str): The message to display in the message box.
    
    Returns:
        bool: Returns True if the user clicks 'Yes', indicating they want to proceed with the process, 
              and False if the user clicks 'No', indicating they do not want to continue.
    """
    window = tk.Tk()
    window.withdraw()  # Hide the main window
    response = messagebox.askyesno(start_title, start_message)
    window.destroy()
    return response



# Función principal para renombrar archivos
def rename_files_locally(df_to_process, directory, error_file_not_found, error_file_already_processed, error_file_already_exists, template_error_not_found, template_error_not_allowed, column_status, proceed_title, proceed_message, cancel_title, cancel_message):
    """
    This function renames files locally within the same source and destination directory.
    It first asks the user if they wish to proceed. If not, the process is canceled.
    
    Args:
        df_to_process (DataFrame): The DataFrame containing original and new file names.
        directory (str): The directory where the files are located.
        error_file_not_found (str): Message for when a file is not found in the directory.
        error_file_already_processed (str): Message for files that were already processed.
        error_file_already_exists (str): Message for files that already exist with the new name.
        template_error_not_found (str): Message for template errors if file not found.
        template_error_not_allowed (str): Message for template errors if permissions are denied.
        column_status (str): Column name to store the processing status.
        proceed_title (str): Title of the prompt asking to proceed.
        proceed_message (str): Message of the prompt asking to proceed.
        cancel_title (str): Title of the message shown if the operation is canceled.
        cancel_message (str): Message shown if the operation is canceled.

    Returns:
        DataFrame: The updated DataFrame with the status column.
    """
    # Ask the user if they want to proceed
    if not ask_to_proceed(proceed_title, proceed_message):
        # If user clicks "No", show a cancellation message and exit
        show_error(cancel_title, cancel_message)
        exit_if_directly_executed()  # Cancel the script

    # List to store the status messages
    statuses = []
    processed_names = set()  # Set to track unique names of 'New Complete Name'

    # Iterate over the DataFrame to rename files and directories
    for index, row in df_to_process.iterrows():
        original_name = row.iloc[3]  # Column 4 = Original file or directory name
        new_name = row.iloc[4]       # Column 5 = New name for the file or directory

        # Full path of the original file or directory
        original_path = os.path.join(directory, original_name)
        # Full path for the new file or directory (in "Modified Name")
        new_path = os.path.join(directory, new_name)

        # Check if the file or directory exists in the source directory
        if not os.path.exists(original_path):
            statuses.append(error_file_not_found)
            continue  # Skip if the original file or directory doesn't exist

        # Check if the new name has already been processed (duplicate in DataFrame)
        if new_name in processed_names:
            statuses.append(error_file_already_processed)
            continue
        
        # Add the new name to the processed set
        processed_names.add(new_name)

        # Check if a file or directory with the new name already exists
        if os.path.exists(new_path):
            # If it already exists, record the status and continue
            statuses.append(error_file_already_exists)
            continue

        # Try renaming the file or directory
        try:
            os.rename(original_path, new_path)
            # Add 'Ok' status if renaming was successful
            statuses.append("Ok")
        except FileNotFoundError:
            statuses.append(template_error_not_found)
        except PermissionError:
            statuses.append(template_error_not_allowed)
        except Exception as e:
            statuses.append(f"Error: {e}")

    # Add the status list to the DataFrame as a new column
    df_to_process[column_status] = statuses

    # Return the updated DataFrame
    return df_to_process


# MAIN MENU FUNCTIONS
def adjust_text(event, *args, margin):
    """
    Adjusts the text wrapping length for given labels based on the available width minus a specified margin.

    Args:
        event: The triggering event, typically a window resize event.
        *args: Variable number of label widgets to adjust.
        margin (int): The margin to subtract from the label width for text wrapping.

    Returns:
        List of labels with updated wraplength configurations (optional).
    """
    updated_labels = []
    for label in args:
        # Adjust the text wrapping length to the available width minus the margin
        new_wraplength = label.winfo_width() - margin
        if new_wraplength > 0:  # Ensure the new wrap length is positive
            label.config(wraplength=new_wraplength)
            updated_labels.append(label)  # Store updated label

    return updated_labels  # Return the list of updated labels



def relative_route_to_file(path_to_folder, file):
    """
    Returns the relative path to a file, accounting for whether the application 
    is running in a PyInstaller bundle or a standard script environment.

    Args:
        path_to_folder (str): The folder path where the file is located.
        file (str): The name of the file.

    Returns:
        str: The complete relative path to the specified file.
    """
    if hasattr(sys, '_MEIPASS'):
        # Running as a PyInstaller bundle, use _MEIPASS to locate files
        route = os.path.join(sys._MEIPASS, path_to_folder, file)
    else:
        # Running as a regular script, use standard relative path
        route = os.path.join(path_to_folder, file)
    
    return route



def open_web_page(*links):
    """
    Attempts to open a list of web links in the default web browser. For each link,
    it checks if the URL is accessible before opening it. Stops at the first successful link.

    Args:
        *links (str): One or more URLs to open.

    Returns:
        None
    """
    for link in links:
        try:
            # Perform a GET request to check if the link is accessible
            response = requests.get(link, timeout=5)
            if response.status_code == 200:
                webbrowser.open(link)
                print(f"Opening: {link}")
                break  # Stop if the link was successfully opened
            else:
                print(f"The link {link} is not available, status code: {response.status_code}")
        except requests.ConnectionError:
            print(f"Connection error while checking {link}. The link might not exist.")
        except requests.Timeout:
            print(f"Timeout expired while checking {link}.")
        except requests.RequestException as e:
            print(f"Error while checking {link}: {e}")
    else:
        print("No links could be opened.")


        

        
        
class DirectExecutionExit(Exception):
    """Custom exception to handle script exit without closing the Tkinter menu."""
    pass

def exit_if_directly_executed():
    """
    Raises an exception that can be caught to stop the script.
    
    Args:
        None
        
    Returns:
        None
    """
    raise DirectExecutionExit("The script was stopped.")




# FUNCTIONS TO EXECUTE SCRIPTS
def execute_script_src(script, language):
    """
    Executes the corresponding script and handles the language.

    Arg:
        script (str): The name of the script to execute.
        language (str): The language parameter to be passed to the script's main function.

    Returns:
        None: This function does not return any value. It only executes the script and handles errors.
    """
    try:
        # Attempt to import and execute the corresponding script
        module = __import__(f'src.{script}', fromlist=['main'])
        module.main(language)  # Calls the main function passing the language

    except DirectExecutionExit as e:
        print(f"The script was stopped: {e}")
    except ImportError as e:
        print(f"Error importing the script '{script}': {e}")
    except AttributeError as e:
        print(f"The script '{script}' does not have a 'main' function: {e}")
    except Exception as e:
        print(f"An error occurred while executing the script '{script}': {e}")




# Functions to execute a specific script
def execute_script0(language):
    """
    Executes the 'a_Create_excel' script with the given language.

    Arg:
        language (str): The language to be passed to the 'a_Create_excel' script.

    Returns:
        None: This function does not return any value. It only calls the execute_script_src function.
    """
    execute_script_src('a_Create_excel', language)


def execute_script1(language):
    """
    Executes the 'b_Modify_files' script with the given language.

    Arg:
        language (str): The language to be passed to the 'b_Modify_files' script.

    Returns:
        None: This function does not return any value. It only calls the execute_script_src function.
    """
    execute_script_src('b_Modify_files', language)



def execute_script2(language):
    """
    Executes the 'c_Unlock_excel' script with the given language.

    Arg:
        language (str): The language to be passed to the 'c_Unlock_excel' script.

    Returns:
        None: This function does not return any value. It only calls the execute_script_src function.
    """
    execute_script_src('c_Unlock_excel', language)

